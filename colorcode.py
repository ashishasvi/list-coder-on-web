import io
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
color_map = {
    "RED": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
    "GREEN": PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid"),
    "YELLOW": PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
}
def apply_color_to_excel(excel_bytes, color_column_name):
    """Applies color formatting based on the 'DateColor' column in Excel before downloading."""

    # Load the Excel file from in-memory BytesIO object
    excel_bytes.seek(0)  # Reset the pointer to the beginning
    wb = load_workbook(excel_bytes)
    ws = wb.active  # Use the first sheet

    # Identify the column index for 'DateColor'
    header_row = [cell.value for cell in ws[1]]  # Read the first row (column headers)
    if color_column_name not in header_row:
        print(f"Column '{color_column_name}' not found in the Excel file.")
        return excel_bytes  # Return without modification

    color_col_index = header_row.index(color_column_name) + 1  # Convert to 1-based index

    # Loop through the 'DateColor' column and apply color formatting
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        color_value = str(row[color_col_index - 1].value).strip().upper()  # Get DateColor value
        if color_value in color_map:
            fill_style = color_map[color_value]  # Get the corresponding color
            for cell in row:  # Apply the color to the entire row
                cell.fill = fill_style
    #####clear contents of columns "Matched" and "DateColor"




    # Get the header row (first row)
    header_row = [cell.value for cell in ws[1]]  

# Find column indexes for "Matched" and "DateColor"
    columns_to_clear = [i + 1 for i, col_name in enumerate(header_row) if col_name in ["Matched", "DateColor"]]

# Clear content of these columns
    for col in columns_to_clear:
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col, max_col=col):
            for cell in row:
                cell.value = None  # Clear cell content            

    # Save the modified file into a new BytesIO object
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)  # Reset pointer before sending
    return output